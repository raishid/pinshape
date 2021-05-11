from openpyxl import load_workbook

excel = load_workbook("Pinshape.xlsx")
hoja = excel.active

class Read_excel:

    @staticmethod
    def Get_data_excel(fila):
        user = hoja.cell(fila, 1).value
        web = hoja.cell(fila, 2).value
        bio = hoja.cell(fila, 3).value
        proxie = hoja.cell(fila, 4).value
        email = hoja.cell(fila, 5).value
        if user is None or bio is None or email is None:
            return False
        else:
            data = {
                "user": user,
                "bio": bio,
                "web": web,
                "proxie": proxie,
                "email": email
            }
            if proxie == None:
                data['proxie'] = None
            return data
    @staticmethod
    def Guardar_link(fila, link):
        hoja.cell(fila, 6).value = link
        excel.save("Pinshape.xlsx")